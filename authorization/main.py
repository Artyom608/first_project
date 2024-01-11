from flask import redirect, send_file
import time
import pymongo
import tempfile
from flask import Flask, request, jsonify
import json
from pymongo import MongoClient
from datetime import datetime, timedelta
import pytz
import jwt
import requests
import random
import string
import os
from typing import Dict, Any, List
from openpyxl import load_workbook
import threading
from flask import render_template, redirect, make_response

app = Flask(__name__)

# Глобальная переменная для проверки, что пользователь дал доступ
authenticating = {'is_done': False, 'code': ''}

# Данные GitHub приложения(OauthAPP)
# ID клиента
CLIENT_ID = "33019ee2f80b715fcc43"
# Секретный ID
CLIENT_SECRET = "f2eedbb81f59afb67fb9458826a99a863ac2b83d"

# Класс для представления данных о клиентах
class Clients:
    def __init__(self, _id, group, username, tg_id, role):
        self.Id = _id
        self.Group = group
        self.Username = username
        self.Tg_id = tg_id
        self.Role = role

# Класс для представления данных о пользователе
class UserData:
    def __init__(self, _id, name):
        self.Id = _id
        self.Name = name

# Функция для добавления клиента в базу данных
def add_client(_id, group, username, tg_id, role):
    # создаём соединение с MongoDB
    client = MongoClient("mongodb+srv://MYDBUSER:123@cluster0.pbupe3g.mongodb.net/?retryWrites=true&w=majority")
    db = client["tg"]
    # обращаемся к коллекции clients из базы tg
    collection = db["clients"]
    # создаём переменную в виде экземпляра класса Clients
    if role == "":
        role = "student" # По умолчанию всем пользователям выдаётся роль "студент"
    current_client = Clients(_id, group, username, tg_id, role)
    # добавляем одиночный документ в коллекцию
    insert_result = collection.insert_one(current_client.__dict__)
    # выводим внутренний ID добавленного документа
    print("Inserted a single document:", insert_result.inserted_id)

# Функция для поиска клиента в базе данных
def find_client(id):
    # создаём соединение с MongoDB
    client = MongoClient("mongodb+srv://MYDBUSER:123@cluster0.pbupe3g.mongodb.net/?retryWrites=true&w=majority")
    db = client["tg"]
    # обращаемся к коллекции clients из базы tg
    collection = db["clients"]
    # создаём фильтр по которому будем искать клиента
    filter_query = {"Id": id}
    # собственно ищем
    result = collection.find_one(filter_query)
    if not result:
        return Clients("", "", "", "", "")
    return Clients(result["Id"], result["Group"], result["Username"], result["Tg_id"], result["Role"])
        
# Функция для обновления данных клиента в базе данных
def update_client(_id, key, value):
    # создаём соединение с MongoDB
    client = MongoClient("mongodb+srv://MYDBUSER:123@cluster0.pbupe3g.mongodb.net/?retryWrites=true&w=majority")
    db = client["tg"]
    # обращаемся к коллекции clients из базы tg
    collection = db["clients"]
    filter_query = {"Id": _id}
    update_query = {"$set": {key: value}}
    collection.update_one(filter_query, update_query)

# Функция для удаления клиента из базы данных по tg_id
def delete_client_by_tg_id(tg_id):
    # создаём соединение с MongoDB
    client = MongoClient("mongodb+srv://MYDBUSER:123@cluster0.pbupe3g.mongodb.net/?retryWrites=true&w=majority")
    db = client["tg"]
    # обращаемся к коллекции clients из базы tg
    collection = db["clients"]
    # Пробуем удалить запись из MongoDB
    filter_query = {"Tg_id": tg_id}
    result = collection.delete_one(filter_query)
    
    if result.deleted_count > 0:  # Если удаление прошло успешно
            return True  # Возвращаем True, если пользователь был успешно удален
    return False  # Возвращаем False, если запись не была найдена или не была удалена


# Функция для проверки, есть ли клиент в базе данных
def is_in_data(code):
    cur_user = find_client(code)
    return bool(cur_user.Id)

# Обработчик для обновления данных клиента через HTTP GET-запрос
@app.route('/update', methods=['GET'])
def update_handler():
    _id = request.args.get("Id") # git_id пользователя
    key = request.args.get("Key") # параметр
    value = request.args.get("Value") # новое значение
    update_client(_id, key, value)
    print("Функция update сработала")
    return "Updated successfully"

# Обработчик для удаления клиента через HTTP GET-запрос
@app.route('/del', methods=['GET'])
def del_handler():
    tg_id = request.args.get("tg_id")

    # Проверяем наличие пользователя с указанным tg_id перед удалением
    if delete_client_by_tg_id(tg_id):
        return "Deleted successfully"
    else:
        return "User not found"

# Обработчик для генерации URL для аутентификации через GitHub
@app.route('/auth', methods=['GET'])
def reg_handler():
    _id = request.args.get("chat_id")
    auth_url = f"https://github.com/login/oauth/authorize?client_id={CLIENT_ID}&state={_id}"
    return auth_url

# Обработчик для добавления клиента через HTTP GET-запрос
@app.route('/add', methods=['GET'])
def add_handler():
    _id = request.args.get("id")
    group = request.args.get("group")
    username = request.args.get("username")
    tg_id = request.args.get("tg_id")
    role = request.args.get("role")  
    add_client(_id, group, username, tg_id, role)
    return "Added successfully"

# Обработчик для поиска клиента и генерации JWT-токена через HTTP GET-запрос
@app.route('/find', methods=['GET'])
def find_handler():
    tg_id = request.args.get("github_id")
    #_id = request.args.get("_id")  # Получаем значение _id из запроса
    user = find_client(tg_id)  # Передаем значение
    token = zacode(user)
    return token

# Функция для генерации JWT-токена на основе данных клиента
def zacode(client):
    #token_expires_at = datetime.now(pytz.UTC) + timedelta(minutes=60)
    token_expires_at = datetime.now(pytz.UTC) + timedelta(minutes=60)
    unix_timestamp = int(token_expires_at.timestamp())
    user = {
        "name": client.Username,
        "id": client.Id,
        "role": client.Role,
        "tg_id": client.Tg_id,
        "expires_at": unix_timestamp, #int(token_expires_at.timestamp())
    }
    token = jwt.encode(user, "123", algorithm="HS256")
    return token

# Функция для обработки запроса администратора
@app.route('/startAdmin', methods=['POST'])
def start_admin_handler():
    data = request.form.get("data")  # Извлекаем данные из запроса
    # Пытаемся декодировать JWT (JSON Web Token), содержащийся в переменной 'data'
    try:
        payload = jwt.decode(data, "123", algorithms=["HS256"])
        # Если поле action == true
        if payload["action"] == True:
            # Извлекаем github_id
            git_id = int(payload["git_id"])
            # Добавляем в словарь sessions
            sessions[git_id] = data
            # Если всё успешно, происходит перенаправление
            return redirect("http://127.0.0.1:8082/admin?token=" + data)
        else:
            return "У Вас нет прав на данное действие"
    except jwt.ExpiredSignatureError:
        return "Истек срок действия токена"
    except jwt.InvalidTokenError:
        return "Неверный токен"

# Функция для входа в режим администратора
@app.route('/enterAdmin', methods=['GET'])
def enter_admin_handler():
    token = request.args.get("token")  # Извлекаем временный код из запроса
    if token:
        if value_exists(sessions, token):
            try:
                # Проверяем токен и извлекаем данные

                # Пробуем декодировать токен

                # Полезные данные, хранящиеся внутри JWT
                payload = jwt.decode(token, "123", algorithms=["HS256"])
                expires_at = int(payload["expires_at"])

                # Проверяем, истекло ли время действия токена
                if expires_at > int(time.time()):
                    # Обновляем время истечения токена
                    new_expiration_time = int(time.time()) + 15 * 60
                    payload["expires_at"] = new_expiration_time

                    # Подписываем обновленный токен
                    updated_token = jwt.encode(payload, "123", algorithm="HS256")

                    # Создаем и устанавливаем куку с обновленным токеном
                    response = make_response(render_template("index.html"))
                    # Создаем HTTP-ответ, включая обновленный токен в виде cookie
                    response.set_cookie("jwt_token", updated_token, expires=new_expiration_time, httponly=True)
                    return response
                else:
                    return "<html><body><h1>Время жизни токена истекло</h1></body></html>"
            except jwt.ExpiredSignatureError:
                return "<html><body><h1>Истек срок действия токена</h1></body></html>"
            except jwt.InvalidTokenError:
                return "<html><body><h1>Неверный токен</h1></body></html>"

        else:
            return "<html><body><h1>Такого пользователя не существует</h1></body></html>"

    else:
        # Генерируем новый токен и отправляем ссылку для авторизации
        token = generate_token(random.randint(2, 5))
        sessions[1] = token
        return f"<html><body><h1><a href='https://web.telegram.org/a/#6609893395'>Перейдите по ссылке для авторизации</a></h1></body></html><html><body><h2>Токен запроса доступа: {token}</h2></body></html>"


# Обработчик для обработки OAuth-авторизации через GitHub
@app.route('/oauth', methods=['POST', 'GET'])
def handle_oauth():
    response_html = "<html><body><h1>Вы НЕ аутентифицированы!</h1></body></html>"

    # Извлекаем параметры "code" и "state" из строки запроса
    code = request.args.get("code")
    tg_id = request.args.get("state")

    if code:
        # Устанавливаем флаг аутентификации, сохраняем код авторизации 
        # в словаре 'authenticating' и обновляем response_html
        authenticating['is_done'] = True
        authenticating['code'] = code
        response_html = "<html><body><h1>Вы аутентифицированы!</h1></body></html>"

        # Извлекаем токен доступа с использованием кода авторизации
        access_token = get_access_token(authenticating['code'])
        user_info = get_user_data(access_token)

        # Проверяем, зарегистрирован ли пользователь с данным идентификатором. 
        # Если нет, он добавляется в базу данных с идентификаторами
        if is_in_data(str(user_info["id"])):
            response_html += "<html><body><h1>Уже зарегистрированы</h1></body></html>"
        else:
            add_client(str(user_info["id"]), "", "", tg_id, "student")
            response_html += "<html><body><h1>Успешно зарегистрированы</h1></body></html>"

        # Формируем URL для подтверждения регистрации
        url = f"http://localhost:8081/register-confirm?chat_id={tg_id}&github_id={user_info['id']}"
        response = requests.get(url)

        if response.status_code == 200:
            print("Все хорошо")

    return response_html

# Функция для выхода из режима администратора
@app.route('/exitAdmin', methods=['GET'])
def exit_admin_handler():
    # Создаем куку с пустым значением и именем "jwt_token"
    cookie = make_response()
    cookie.set_cookie("jwt_token", "", expires=0, httponly=True)

    # Перенаправляем пользователя на текущую страницу
    return redirect(request.url, code=302)

# Функция для обновления данных студента
@app.route('/updateStudent', methods=['POST'])
def update_student_handler():
    # Извлекаем токен из куки запроса. 
    token_cookie = request.cookies.get("jwt_token")

    # Если токен отсутствует, возвращаем сообщение об ошибке
    if not token_cookie:
        return render_template("error.html", message="Доступ запрещен")

    # Пытаемся декодировать токен
    try:
        payload = jwt.decode(token_cookie, "123", algorithms=["HS256"])
        expires_at = int(payload["expires_at"])

        # Проверяем срок действия и разрешение на выполнение действия
        if expires_at > int(time.time()) and payload["action"]:
            # Извлекаем данные из POST-запроса
            user_id = int(request.form.get("git_id"))
            username = request.form.get("username")
            group = request.form.get("group")
            role = request.form.get("role")

            # Подключение к базе данных
            client_options = pymongo.MongoClient("mongodb+srv://MYDBUSER:123@cluster0.pbupe3g.mongodb.net/?retryWrites=true&w=majority")
            client = pymongo.MongoClient(client_options)

            # Выполняем обновление данных студента
            try:
                db = client["TGbot"]
                collection = db["user"]

                filter_query = {"github_id": user_id}
                update_query = {"$set": {"role": role, "about.username": username, "about.group": group}}

                result = collection.update_one(filter_query, update_query)

                if result.matched_count != 0:
                    print("Matched and replaced an existing document")
                elif result.upserted_id is not None:
                    print("Inserted a new document with ID", result.upserted_id)

            except Exception as e:
                print(f"An error occurred: {e}")
                return render_template("error.html", message="Internal Server Error")

            finally:
                client.close()

            # После успешного обновления данных в базе данных обновляется срок действия токена 
            new_expiration_time = int(time.time()) + 15 * 60
            payload["expires_at"] = new_expiration_time

            # Создаём новый токен
            new_token = jwt.encode(payload, "123", algorithm="HS256")
            response = make_response(render_template("index.html"))
            response.set_cookie("jwt_token", new_token, expires=new_expiration_time, httponly=True)

            return response

    # Если условия проверки токена не выполняются, возвращаем ошибки
        else:
            return render_template("error.html", message="Время жизни токена истекло или у вас нет прав на данное действие")

    except jwt.ExpiredSignatureError:
        return render_template("error.html", message="Истек срок действия токена")

    except jwt.InvalidTokenError:
        return render_template("error.html", message="Неверный токен")

# Загрузка расписания в админке
@app.route('/uploadSchedule', methods=['POST'])
def upload_schedule_handler():
    # Извлекаем токен из куки запроса
    token_cookie = request.cookies.get("jwt_token")

    if not token_cookie:
        return render_template("error.html", message="Доступ запрещен")

    # Пытаемся декодировать токен
    try:
        payload = jwt.decode(token_cookie, "123", algorithms=["HS256"])
        expires_at = int(payload["expires_at"])

        if expires_at > int(time.time()) and payload["action"]:
            # Далее извлекаем файл из POST-запроса, сохраняем временно на сервере
            file = request.files.get("file")

            if file:
                try:
                    file.save(file.filename)
                    # Вызываем функцию parse_excel_to_json для обработки файла. 
                    parse_excel_to_json(file.filename)

                except Exception as e:
                    print(f"An error occurred: {e}")
                    return render_template("error.html", message="Ошибка при обработке файла")

                finally:
                    os.remove(file.filename)

                # Если файл успешно обработан, обновляем срок действия токена
                new_expiration_time = int(time.time()) + 15 * 60
                payload["expires_at"] = new_expiration_time

                # Создаем новый токен
                new_token = jwt.encode(payload, "123", algorithm="HS256")
                # Устанавливаем в куки ответа перенаправление на index.html
                response = make_response(render_template("index.html"))
                response.set_cookie("jwt_token", new_token, expires=new_expiration_time, httponly=True)

                return response

        # Если условия проверки не выполняются, возвращаем ошибки
            else:
                return render_template("error.html", message="Файл не был передан")

        else:
            return render_template("error.html", message="Время жизни токена истекло или у вас нет прав на данное действие")

    except jwt.ExpiredSignatureError:
        return render_template("error.html", message="Истек срок действия токена")

    except jwt.InvalidTokenError:
        return render_template("error.html", message="Неверный токен")

# Функция для получения access_token через OAuth-авторизацию с GitHub(для получения данных пользователя с GitHub)
def get_access_token(code):
    # URL, на который будет отправлен POST-запрос для получения access_token.
    request_url = "https://github.com/login/oauth/access_token"
    # form - словарь, содержащий параметры, которые нужно отправить вместе с запросом. 
    form = {
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "code": code,
    }
    # Заголовки HTTP для запроса.
    headers = {"Accept": "application/json"}
    # POST - запрос
    response = requests.post(request_url, data=form, headers=headers)
    # JSON - ответ
    response_json = response.json()
    return response_json.get("access_token", None)

# Обновление расписания из облака
def update_schedule(file_url):
    try:
        # Запрос к указанному файлу
        response = requests.get(file_url)

        if response.status_code == 200:
            data = response.content

            # Создаём временный файл с расширением ".xlsx" с использованием модуля tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
                temp_file.write(data)
                # Путь к файлу
                temp_file_path = temp_file.name

                parse_excel_to_json(temp_file_path)

            return temp_file_path

        else:
            print(f"Server error: {response.status_code}")
            return None

    except Exception as e:
        print(f"An error occurred: {e}")
        return None

# Используется для отправки файла клиенту
@app.route('/downloadSchedule', methods=['GET'])
def download_schedule_handler():
    file_url = "https://example.com/schedule.xlsx"  
    temp_file_path = update_schedule(file_url)

    if temp_file_path:
        return send_file(temp_file_path, as_attachment=True)

    return render_template("error.html", message="Ошибка при обновлении расписания из облака")

# Функция для получения данных пользователя с GitHub
def get_user_data(access_token):
    request_url = "https://api.github.com/user"
    headers = {"Authorization": f"Bearer {access_token}"}
    response = requests.get(request_url, headers=headers)
    return response.json() if response.status_code == 200 else None

# Парсер расписания
def parse_excel_to_json(filename: str) -> None:
    wb = load_workbook(filename)
    sheet = wb.active

    # data - основной массив данных,
    data = []
    # item - словарь для хранения данных о четных/нечетных неделях, группах и расписании.
    item = {}

    # Четная/нечетная неделя
    # итерация по ячейкам в первой строке 
    for key in sheet[1]:
        # пустой словарь внутри словаря item с ключом, соответствующим значению в ячейке первой строки.
        item[key.value] = {}
    # Удаляются возможные пустые значения в словаре 
    item.pop(None, None)

    # Группы
    for group in item:
        # ссылка на вложенный словарь для текущей группы
        cur = item[group]
        # по ячейкам второй строки листа Excel
        for key in sheet[2]:
            # пустой словарь внутри словаря cur с ключом, соответствующим значению в ячейке второй строки
            cur[key.value] = {}

            # Дни недели
            days = cur[key.value]
            #итерация по строкам, начиная с третьей (индекс 2) до последней строки, с шагом 6
            for ind in range(2, sheet.max_row, 6):
                # если текущая строка соответствует строке, представляющей дни недели
                if sheet[ind].row == 2:
                    # Итерация по дням недели в текущей строке.         
                    for key in sheet[ind]:
                        # пустой словарь внутри словаря days 
                        days[key.value] = {}

                        # Предметы
                        classes = days[key.value]
                        # Итерация по строкам, представляющим предметы для текущего дня недели.
                        for row in range(ind + 1, ind + 7):
                            # Создается запись в словаре classes
                            classes[sheet[row][1].value] = {
                                "Предмет": sheet[row][2].value,
                                "Аудитория": sheet[row][3].value,
                                "Комментарий": "",
                            }
                    # Удаляются возможные пустые значения в словаре 
                    days.pop(None, None)

    data.append(item)

    # Парсим в JSON
    json_data = json.dumps(data)

    # Отправляем POST-запрос
    url = "http://localhost:8080/update-schedule"
    # Определяются заголовки HTTP для запроса.
    headers = {"Content-Type": "application/json"}
    try:
        # выполняется POST-запрос к указанному URL. 
        response = requests.post(url, data=json_data, headers=headers)
        # проверяется, был ли успешным запрос
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"Ошибка при отправке запроса: {e}")

    return json_data  # возвращаем данные

# Конвертация интерфейса в словарь
def convert_interface_to_map(i: Any) -> Dict[str, Any]:
    # Проверяем, что интерфейс является типом словаря
    if isinstance(i, dict):
        return i
    return {}

# Проверка, существует ли значение в словаре
def value_exists(mymap: Dict[int, str], value: str) -> bool:
    return value in mymap.values()

# Генерация случайного токена запроса доступа
def generate_token(length: int) -> str:
    # Допустимые символы для генерации токена
    letters = string.ascii_letters + string.digits

    # Создание токена заданной длины
    token = ''.join(random.choice(letters) for _ in range(length))
    return token

def check_schedule_changed():
    # Настройка периодичности обновления расписания (например, раз в сутки)
    update_interval = 10 * 3600  # в секундах

    # URL, где хранится файл с расписанием
    file_url = "https://ssou.ru/Pars.xlsx"

    # Запуск бесконечного цикла обновления в отдельном потоке
    def update_schedule_periodically():
        while True:
            time.sleep(update_interval)
            try:
                update_schedule(file_url)
            except Exception as e:
                print(f"Ошибка при обновлении расписания: {e}")

    # Создание и запуск потока
    update_thread = threading.Thread(target=update_schedule_periodically)
    update_thread.start()

@app.route('/get_all_users', methods=['GET'])
def get_all_users_handler():
    # создаём соединение с MongoDB
    client = MongoClient("mongodb+srv://MYDBUSER:123@cluster0.pbupe3g.mongodb.net/?retryWrites=true&w=majority")
    db = client["tg"]
    # обращаемся к коллекции clients из базы tg
    collection = db["clients"]
    
    # извлекаем все документы из коллекции
    all_users = list(collection.find())
    
    # создаём список для хранения данных о пользователях
    users_data = []
    
    # формируем список с данными о каждом пользователе
    for user in all_users:
        user_data = {
            "Id": user["Id"],
            "Username": user["Username"],
            "Tg_id": user["Tg_id"],
            "Role": user["Role"],
            "Group": user["Group"]
        }
        users_data.append(user_data)
    # возвращаем JSON с данными о пользователях
    print(users_data)
    return jsonify(users_data)

if __name__ == '__main__':
    app.run(port=8080, debug=True)

# Добавляем вызов функции в код
if __name__ == '__main__':
    check_schedule_changed()
